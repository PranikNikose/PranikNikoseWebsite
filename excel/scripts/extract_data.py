"""
Regenerates the `DATA` array embedded in index.html from the source Excel
workbook. This is the ONLY place question data should ever be edited — if
you need to fix/add a question, edit the .xlsx, then re-run this script.
Never hand-edit the DATA array inside index.html directly; it's a build
artifact, not a source file, and will be silently overwritten next time
this script runs. See docs/RULES.md section "Excel -> index.html sync workflow".

Usage:
    python scripts/extract_data.py

Fully standalone -- no other tool is needed. This auto-detects the Excel
workbook to read (whatever single .xlsx file sits in the project root -- see
find_workbook() below, no filename to configure), prints a per-sheet row
count, and writes the result directly into index.html's
`var DATA = [ ... ];` array (a full overwrite of just that array --
nothing else in index.html is touched). Rich text (bold/italic/underline/
color) is preserved automatically via openpyxl's rich_text mode, exactly
as it's formatted in the workbook.

Before overwriting, the previous index.html is backed up to
index.html.bak -- if a run looks wrong, restore it by copying that file
back over index.html. If the DATA array can't be found in index.html at
all, the script raises an error and writes nothing (never guesses where
to put the data).

To change WHICH sheets get pulled, or to switch between a sampled/testing
row cap and the full dataset, edit the SHEETS/ROWS_PER_SHEET constants
below before running -- see docs/SYNC_EXCEL.md for guidance on what values to
use for common requests ("sync everything", "just these 3 sheets", etc).
"""

import openpyxl, json, html as htmlmod, re, os, glob
from openpyxl.cell.rich_text import CellRichText, TextBlock

# This script lives in scripts/ -- resolve everything relative to the
# PROJECT ROOT (this file's parent's parent), not the current working
# directory, so `python extract_data.py` / `python scripts/extract_data.py`
# both work correctly no matter where the caller's shell cwd happens to be.
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def find_workbook():
    """There must be exactly one .xlsx file in the project root -- no
    filename is hardcoded/configured, it's just whatever workbook is
    sitting there. Excel's own temp lock files (~$actualname.xlsx,
    created while the real file is open in Excel) are ignored, not
    counted as a second workbook. Raises if there's zero or more than one
    real .xlsx."""
    pattern = os.path.join(PROJECT_ROOT, '*.xlsx')
    candidates = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith('~$')]
    if len(candidates) == 0:
        raise RuntimeError(
            'No .xlsx workbook found in ' + PROJECT_ROOT + '. Place exactly '
            'one Excel file there and re-run.'
        )
    if len(candidates) > 1:
        raise RuntimeError(
            'Found ' + str(len(candidates)) + ' .xlsx files in ' + PROJECT_ROOT +
            ' -- there must be exactly one:\n' +
            '\n'.join('  - ' + os.path.basename(f) for f in candidates) +
            '\nRemove/move the extras, keeping only the one workbook to sync from.'
        )
    return candidates[0]


WORKBOOK = find_workbook()
SHEETS = ['CoreJava', 'SpringBoot', 'Microservices']  # HR excluded: personal PII, see docs/RULES.md

# Every real, non-HR sheet in the workbook -- used by add_sheets.py's
# "--all" mode and by smart_sync.py. Keep in sync with docs/RULES.md section 3
# if the workbook's tabs are ever renamed/added/removed.
ALL_KNOWN_SHEETS = ['Coding', 'CoreJava', 'Adv. Java', 'Junit', 'Maven', 'JDBC',
                     'Spring Frmwrk', 'Hibernate', 'Spring Sec.', 'SpringBoot',
                     'Microservices', 'DSA', 'API', 'RDBMS', 'Design Pattern',
                     'AI', 'React']

# Full row count for these 3 user-selected sheets (docs/SYNC_EXCEL.md sec 2b).
ROWS_PER_SHEET = None

OR_SEPARATOR_HTML = ' <em>or</em> '
OR_SEPARATOR_PLAIN = ' or '

wb = openpyxl.load_workbook(WORKBOOK, rich_text=True)


def find_col(header, *needles):
    for i, h in enumerate(header):
        if h is None:
            continue
        hl = str(h).strip().lower()
        for n in needles:
            if hl.startswith(n):
                return i + 1
    return None


def plain(v):
    if v is None:
        return ''
    if isinstance(v, CellRichText):
        return str(v).strip()
    return str(v).strip()


def font_color(font):
    try:
        c = font.color
        if c is not None and c.type == 'rgb' and c.rgb and len(c.rgb) == 8:
            argb = c.rgb
            if argb[:2] != '00':
                return '#' + argb[2:]
    except Exception:
        pass
    return None


def lines_with_runs(value):
    if value is None:
        return [[]]
    items = value if isinstance(value, CellRichText) else [value]
    lines = [[]]
    for item in items:
        if isinstance(item, TextBlock):
            text, font = item.text, item.font
        else:
            text, font = str(item), None
        parts = text.split('\n')
        for idx, part in enumerate(parts):
            if part:
                lines[-1].append((part, font))
            if idx < len(parts) - 1:
                lines.append([])
    return lines


def run_to_html(text, font):
    esc = htmlmod.escape(text)
    if font is None:
        return esc
    if getattr(font, 'u', None):
        esc = '<u>' + esc + '</u>'
    if getattr(font, 'i', None):
        esc = '<em>' + esc + '</em>'
    if getattr(font, 'b', None):
        esc = '<strong>' + esc + '</strong>'
    color = font_color(font)
    if color:
        esc = '<span style="color:' + color + '">' + esc + '</span>'
    return esc


def line_to_html(line):
    return ''.join(run_to_html(t, f) for t, f in line if t)


def line_to_plain(line):
    return ''.join(t for t, f in line if t)


def cell_lines(cell_value):
    """Return list of (html, plain) per non-empty line -- used for the
    Questions cell, where each line is a separate phrasing (see
    docs/RULES.md: 'one Excel row = one DATA entry, join phrasings with or,
    keep questionParts as the un-joined list')."""
    lines = lines_with_runs(cell_value)
    out = []
    for line in lines:
        p = line_to_plain(line).strip()
        if not p:
            continue
        out.append((line_to_html(line).strip(), p))
    return out


def normalize_level(label):
    """Fold known phrasing/case variants of the same real level in the
    source Excel into one canonical string, so the Level filter/Interview
    Mode chip list doesn't show near-duplicates as separate values (e.g.
    'Mid-Level (3-5 Years)' and 'Mid (3-5 Years)' are the same level, just
    typed differently across rows/sheets). Only folds variants that are
    unambiguously the same thing (a '-Level' suffix, or 'years'/'year'
    case) -- doesn't touch anything else, so a genuinely different label
    stays distinct. See docs/RULES.md section 3 / section 7 (Interview Mode)."""
    if not label:
        return label
    s = re.sub(r'-Level\b', '', label, flags=re.IGNORECASE)
    s = re.sub(r'\byears\b', 'Years', s, flags=re.IGNORECASE)
    s = re.sub(r'\byear\b', 'Year', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# Years-of-experience bands map to one fixed level name, per the user's
# explicit rule: if the Years column is filled in, that number range IS
# the level -- it overrides whatever text happened to be in the Excel's
# Level column for that row (including mismatched/wrong values like
# "Fresher" paired with "3-5 Years", or a generic "General" fallback that
# actually had real years attached). A years value that falls BETWEEN two
# of these bands (not an exact match to one) still resolves to a band --
# picked by which band contains the years value's midpoint -- rather than
# falling back to "General". Only a years value with no parseable number
# at all, or completely outside 0-10, falls through to "General".
LEVEL_BANDS = [
    (0, 1, 'Fresher'),
    (1, 3, 'Junior to Mid'),
    (3, 5, 'Mid'),
    (5, 7, 'Senior'),
    (7, 10, 'Lead / Architect'),
]


def canonical_level_for_years(years):
    if not years:
        return None
    nums = [float(n) for n in re.findall(r'\d+(?:\.\d+)?', years)]
    if not nums:
        return None
    lo = nums[0]
    hi = nums[1] if len(nums) > 1 else nums[0]
    mid = (lo + hi) / 2
    for band_lo, band_hi, name in LEVEL_BANDS:
        if band_lo <= mid <= band_hi:
            return name
    return None


def answer_html_and_plain(cell_value):
    lines = lines_with_runs(cell_value)
    html_lines = [line_to_html(l) for l in lines]
    plain_lines = [line_to_plain(l) for l in lines]
    return '\n'.join(html_lines).strip(), '\n'.join(plain_lines).strip()


def extract(sheets=None, rows_per_sheet=None):
    """sheets/rows_per_sheet default to the module-level SHEETS/
    ROWS_PER_SHEET constants (used when running this script directly);
    other scripts (add_sheets.py, smart_sync.py) pass their own so they
    don't have to mutate this module's globals to control what gets
    pulled."""
    if sheets is None:
        sheets = SHEETS
    if rows_per_sheet is None:
        rows_per_sheet = ROWS_PER_SHEET

    all_questions = []
    per_sheet_counts = {}

    for sheet in sheets:
        ws = wb[sheet]
        header = [plain(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]

        # Column layout is NOT uniform across sheets (e.g. Coding has no
        # level/years columns) -- map by header name, never assume position.
        col_srno = find_col(header, 'sr')
        col_level = find_col(header, 'level')
        col_years = find_col(header, 'years')
        col_topic = find_col(header, 'topic')
        col_questions = find_col(header, 'question')
        col_answers = find_col(header, 'answer')
        col_priority = find_col(header, 'priority', 'proority')

        taken = 0
        ordinal = 0
        for row in range(2, ws.max_row + 1):
            if rows_per_sheet is not None and taken >= rows_per_sheet:
                break
            q_cell = ws.cell(row=row, column=col_questions).value if col_questions else None
            q_lines = cell_lines(q_cell)  # list of (html, plain) per phrasing
            if not q_lines:
                continue

            ordinal += 1
            srNo = plain(ws.cell(row=row, column=col_srno).value) if col_srno else ''
            if not srNo:
                srNo = str(ordinal)  # source left it blank; use row position within sheet instead

            level = plain(ws.cell(row=row, column=col_level).value) if col_level else ''
            years = plain(ws.cell(row=row, column=col_years).value) if col_years else ''
            topic = plain(ws.cell(row=row, column=col_topic).value) if col_topic else ''
            priority = plain(ws.cell(row=row, column=col_priority).value) if col_priority else ''

            # Level is now driven entirely by Years, not the Excel's Level
            # column text: a years value that resolves to a band uses that
            # band's name; anything else (years blank, or outside 0-10)
            # is "General" -- the original Level cell text is not used as
            # a fallback anymore (per the user's explicit rule).
            level = canonical_level_for_years(years) or 'General'
            topic = topic or sheet
            levelLabel = normalize_level(level + (' (' + years + ')' if years else ''))

            answer_cell = ws.cell(row=row, column=col_answers).value if col_answers else None
            answer_html, answer_plain = answer_html_and_plain(answer_cell)

            question_html = OR_SEPARATOR_HTML.join(h for h, p in q_lines)
            question_plain = OR_SEPARATOR_PLAIN.join(p for h, p in q_lines)
            question_parts = [h for h, p in q_lines]

            # One Excel row = one DATA entry -- never split a multi-line
            # Questions cell into separate entries (see docs/RULES.md section 3).
            all_questions.append({
                'srNo': srNo,
                'sheet': sheet,
                'category': topic,
                'level': levelLabel,
                'question': question_html,
                'questionPlain': question_plain,
                'questionParts': question_parts,
                'answer': answer_html,
                'answerPlain': answer_plain,
                'priority': priority
            })
            taken += 1
        per_sheet_counts[sheet] = taken

    return all_questions, per_sheet_counts


INDEX_HTML = os.path.join(PROJECT_ROOT, 'index.html')
DATA_ARRAY_RE = re.compile(r'var DATA = (\[[\s\S]*?\]);\n')


def to_js_array_literal(questions):
    js = json.dumps(questions, ensure_ascii=False)
    # </script (or any </) inside the data would prematurely close the
    # <script> tag it's embedded in -- escape every "</" as "<\/", which
    # is valid inside a JS string/array literal and invisible once parsed.
    return js.replace('</', '<\\/')


def read_current_data(html_path=INDEX_HTML):
    """Reverse of to_js_array_literal() -- reads the `DATA` array that's
    CURRENTLY embedded in index.html back out as a Python list of dicts.
    Used by add_sheets.py (to see which sheets are already synced) and
    smart_sync.py (to diff against). Raises if the array can't be found."""
    with open(html_path, encoding='utf-8') as f:
        html = f.read()
    m = DATA_ARRAY_RE.search(html)
    if not m:
        raise RuntimeError("Could not find 'var DATA = [ ... ];' in " + html_path)
    return json.loads(m.group(1).replace('<\\/', '</'))


def splice_into_index_html(questions, html_path=INDEX_HTML):
    """Full overwrite of index.html's `var DATA = [ ... ];` array -- the
    one and only thing this touches. Backs up the previous index.html to
    `<html_path>.bak` first, so a bad run can be undone by copying that
    back over index.html. Raises (and writes nothing) if the DATA array
    can't be found, rather than guessing where to put it."""
    with open(html_path, encoding='utf-8') as f:
        current_html = f.read()

    m = DATA_ARRAY_RE.search(current_html)
    if not m:
        raise RuntimeError(
            "Could not find 'var DATA = [ ... ];' in " + html_path +
            " -- aborting without writing anything."
        )

    new_block = 'var DATA = ' + to_js_array_literal(questions) + ';\n'
    new_html = current_html[:m.start()] + new_block + current_html[m.end():]

    backup_path = html_path + '.bak'
    with open(backup_path, 'w', encoding='utf-8') as f:
        f.write(current_html)

    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(new_html)

    return backup_path


if __name__ == '__main__':
    questions, counts = extract()
    print("rows taken per sheet:")
    print(json.dumps(counts, indent=2))
    print("total questions:", len(questions))

    debug_json_path = os.path.join(PROJECT_ROOT, 'extracted_questions.json')
    with open(debug_json_path, 'w', encoding='utf-8') as f:
        json.dump(questions, f, ensure_ascii=False, indent=2)

    backup_path = splice_into_index_html(questions)

    # Splice succeeded (an exception above would have stopped this line
    # from running) -- the JSON was only ever a debugging intermediate,
    # so clean it up now rather than leaving it to accumulate.
    os.remove(debug_json_path)

    print("\nSpliced " + str(len(questions)) + " questions into " + INDEX_HTML + "'s DATA array.")
    print("Previous version backed up to " + backup_path + " -- if this run")
    print("looks wrong, restore it by copying that file back over " + INDEX_HTML + ".")
